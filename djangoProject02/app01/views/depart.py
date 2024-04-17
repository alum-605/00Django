from django.shortcuts import render, redirect, HttpResponse

from app01.models import Department
from openpyxl import load_workbook

"""部门管理"""


def depart_list(request):
    """部门列表"""

    models = Department.objects.all()
    # print(models)
    return render(request, "depart_list.html", {"models": models})


def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, "depart_add.html")
    name = request.POST.get("DepartName")
    Department.objects.create(title=name)
    return redirect("/depart/list")


def depart_edit(request, nid):
    """修改部门"""
    if request.method == "GET":
        row_object = Department.objects.filter(id=nid).first()
        # print(row_object.title)
        return render(request, "depart_edit.html", {"row_object": row_object})
    newname = request.POST.get("DepartName")
    Department.objects.filter(id=nid).update(title=newname)
    return redirect("/depart/list/")


def depart_delete(request, nid):
    """删除部门"""
    Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")


def depart_multi(request):
    """批量上传--excel"""
    # 获取文件对象
    fiel_obj = request.FILES.get("exc")
    # 对象传递给openpyxl， 由其读取文件内容---->得sheet
    work_book_obj = load_workbook(fiel_obj)
    sheet = work_book_obj.worksheets[0]

    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        department = row[0].value
        exists = Department.objects.filter(title=department).exists()
        if not exists:
            Department.objects.create(title=department)
        print(department + "already exists")
    return redirect("/depart/list/")
